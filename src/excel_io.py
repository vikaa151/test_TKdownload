"""Excel 导入导出（openpyxl）。

- 账号/作品一键导出为 xlsx
- 账号批量导入模板 + 读取
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .constants import (
    ACCOUNT_EXPORT_FIELDS, ACCOUNT_IMPORT_FIELDS, WORK_EXPORT_FIELDS,
)
from .models import Account, Work


def _bool(v) -> str:
    return "是" if v else "否"


def _unbool(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip() in ("是", "1", "true", "True", "是")


def export_accounts_xlsx(path: str | Path, accounts: Iterable[Account]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "账号"
    ws.append([label for _, label in ACCOUNT_EXPORT_FIELDS])
    for acc in accounts:
        ws.append([
            acc.nickname, acc.douyin_id, acc.uid, acc.share_url, acc.homepage_url,
            _bool(acc.monitor), _bool(acc.auto_download), acc.download_dir,
            acc.remark, acc.last_update, acc.created_at,
        ])
    wb.save(str(path))


def export_works_xlsx(path: str | Path, works: Iterable[Work]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "作品"
    ws.append([label for _, label in WORK_EXPORT_FIELDS])
    for w in works:
        ws.append([
            w.remark, w.nickname, w.douyin_id, w.publish_time, w.work_type,
            w.description, w.url, w.download_time, w.local_path, w.file_name,
        ])
    wb.save(str(path))


def write_account_template(path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "账号导入模板"
    ws.append([label for _, label in ACCOUNT_IMPORT_FIELDS])
    # 示例行
    ws.append(["示例昵称", "douyin_id_示例", "https://v.douyin.com/xxxx/", "示例备注", "否", "否"])
    wb.save(str(path))


def import_accounts_xlsx(path: str | Path) -> list[dict]:
    """读取账号导入 xlsx，返回规范化字典列表。"""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    label_to_key = {label: key for key, label in ACCOUNT_IMPORT_FIELDS}
    out = []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        d = {}
        for i, cell in enumerate(r):
            label = header[i] if i < len(header) else ""
            key = label_to_key.get(label)
            if key:
                d[key] = "" if cell is None else str(cell).strip()
        if d.get("nickname") or d.get("douyin_id") or d.get("share_url"):
            out.append(d)
    wb.close()
    return out
