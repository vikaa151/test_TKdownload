from src.excel_io import (
    export_accounts_xlsx, export_works_xlsx, write_account_template,
    import_accounts_xlsx,
)
from src.models import Account, Work
from openpyxl import load_workbook


def test_export_accounts(tmp_path):
    accs = [Account(nickname="张三", douyin_id="z", monitor=True, auto_download=False,
                    share_url="https://v.douyin.com/x/")]
    p = tmp_path / "acc.xlsx"
    export_accounts_xlsx(p, accs)
    wb = load_workbook(p)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "昵称"
    assert rows[1][0] == "张三"
    assert rows[1][5] == "是"   # 是否监控
    assert rows[1][6] == "否"   # 自动下载


def test_template_and_import(tmp_path):
    tpl = tmp_path / "tpl.xlsx"
    write_account_template(tpl)
    imported = import_accounts_xlsx(tpl)
    # 示例行应被读取
    assert any(d.get("nickname") == "示例昵称" for d in imported)


def test_export_works(tmp_path):
    works = [Work(remark="r", nickname="n", douyin_id="d", publish_time="2024",
                  work_type="视频", description="desc", url="u", file_name="f.mp4")]
    p = tmp_path / "w.xlsx"
    export_works_xlsx(p, works)
    wb = load_workbook(p)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[1][0] == "r"
    assert rows[1][4] == "视频"
